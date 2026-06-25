from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from tkinter import filedialog


class ExcelExporter:

    def __init__(self):
        pass

    # ======================================================
    # Save File Dialog
    # ======================================================

    def get_save_path(self):

        return filedialog.asksaveasfilename(
            title="Export Results",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel Workbook", "*.xlsx")
            ]
        )

    # ======================================================
    # Export Data
    # ======================================================

    def export(
        self,
        headers,
        rows,
        file_path
    ):

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Query Results"

        # --------------------------------------------------
        # Headers
        # --------------------------------------------------

        header_font = Font(
            bold=True
        )

        header_fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3"
        )

        for column_index, header in enumerate(
            headers,
            start=1
        ):

            cell = worksheet.cell(
                row=1,
                column=column_index
            )

            cell.value = header
            cell.font = header_font
            cell.fill = header_fill

        # --------------------------------------------------
        # Data
        # --------------------------------------------------

        for row_index, row_data in enumerate(
            rows,
            start=2
        ):

            for column_index, value in enumerate(
                row_data,
                start=1
            ):

                worksheet.cell(
                    row=row_index,
                    column=column_index
                ).value = value

        # --------------------------------------------------
        # Auto Fit Columns
        # --------------------------------------------------

        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                try:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except Exception:
                    pass

            adjusted_width = max_length + 4

            worksheet.column_dimensions[
                column_letter
            ].width = adjusted_width

        # --------------------------------------------------
        # Freeze Header Row
        # --------------------------------------------------

        worksheet.freeze_panes = "A2"

        # --------------------------------------------------
        # Save
        # --------------------------------------------------

        workbook.save(file_path)

        return file_path

    # ======================================================
    # Export Using Dialog
    # ======================================================

    def export_with_dialog(
        self,
        headers,
        rows
    ):

        file_path = self.get_save_path()

        if not file_path:
            return None

        return self.export(
            headers,
            rows,
            file_path
        )

    # ======================================================
    # Export From Results Panel
    # ======================================================

    def export_from_results_panel(
        self,
        results_panel
    ):

        headers = results_panel.sheet.headers()

        rows = results_panel.sheet.get_sheet_data()

        return self.export_with_dialog(
            headers,
            rows
        )